"""
Process DOJO - Enhanced Admin Configuration
Clean design, comprehensive guide, and styled Excel exports
"""
from django.contrib import admin
from django.utils.html import format_html
from django.db.models import Count, Avg, Q
from django.urls import path
from django.shortcuts import render
from django.http import HttpResponse
from datetime import datetime, timedelta
from django.utils import timezone

# Excel export
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Import your models here (adjust the import path as needed)
from .models import (
    EmployeeProfile, Unit, Line, Operation, TrainingVideo,
    VideoCompletion, MCQTest, Question, TestAttempt,
    SavedAnswer, LoginSession
)


# ============================================================================
# CONTENT MANAGEMENT - Single consolidated view for training content
# ============================================================================

class LineInline(admin.TabularInline):
    """Inline editing for lines within units"""
    model = Line
    extra = 1
    fields = ['name', 'description', 'is_active']
    show_change_link = True


class OperationInline(admin.TabularInline):
    """Inline editing for operations within lines"""
    model = Operation
    extra = 1
    fields = ['name', 'operation_code', 'is_ctq_station', 'order', 'is_active']
    show_change_link = True


class QuestionInline(admin.StackedInline):
    """Stacked inline for better question editing"""
    model = Question
    extra = 3
    fields = [
        'ordering',
        'question_text',
        ('option_a', 'option_b'),
        ('option_c', 'option_d'),
        'correct_answer',
        'marks',
        'explanation'
    ]


class MCQTestInline(admin.StackedInline):
    """MCQ test inline directly in video admin"""
    model = MCQTest
    extra = 0
    fields = ['title', 'description', 'passing_score', 'time_limit_minutes', 'is_active']
    show_change_link = True


@admin.register(Unit)
class UnitAdmin(admin.ModelAdmin):
    """Manage Units with inline Lines"""
    list_display = ['name', 'line_count', 'operation_count', 'video_count', 'is_active']
    list_filter = ['is_active']
    search_fields = ['name', 'description']
    inlines = [LineInline]
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.annotate(
            _line_count=Count('lines', distinct=True),
            _operation_count=Count('lines__operations', distinct=True),
            _video_count=Count('lines__operations__videos', distinct=True)
        )
    
    def line_count(self, obj):
        return obj._line_count
    line_count.short_description = 'Lines'
    line_count.admin_order_field = '_line_count'
    
    def operation_count(self, obj):
        return obj._operation_count
    operation_count.short_description = 'Operations'
    operation_count.admin_order_field = '_operation_count'
    
    def video_count(self, obj):
        return obj._video_count
    video_count.short_description = 'Videos'
    video_count.admin_order_field = '_video_count'


@admin.register(Line)
class LineAdmin(admin.ModelAdmin):
    """Manage Lines with inline Operations"""
    list_display = ['name', 'unit', 'operation_count', 'video_count', 'is_active']
    list_filter = ['unit', 'is_active']
    search_fields = ['name', 'unit__name']
    inlines = [OperationInline]
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('unit').annotate(
            _operation_count=Count('operations', distinct=True),
            _video_count=Count('operations__videos', distinct=True)
        )
    
    def operation_count(self, obj):
        return obj._operation_count
    operation_count.short_description = 'Operations'
    
    def video_count(self, obj):
        return obj._video_count
    video_count.short_description = 'Videos'


@admin.register(Operation)
class OperationAdmin(admin.ModelAdmin):
    """Simplified operation view"""
    list_display = ['name', 'operation_code', 'line', 'unit_name', 'is_ctq_station', 'video_count', 'is_active']
    list_filter = ['line__unit', 'line', 'is_ctq_station', 'is_active']
    search_fields = ['name', 'operation_code', 'line__name']
    list_editable = ['is_active']
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('line__unit').annotate(
            _video_count=Count('videos')
        )
    
    def unit_name(self, obj):
        return obj.line.unit.name
    unit_name.short_description = 'Unit'
    
    def video_count(self, obj):
        return obj._video_count
    video_count.short_description = 'Videos'


@admin.register(TrainingVideo)
class TrainingVideoAdmin(admin.ModelAdmin):
    """All-in-one video and test management"""
    list_display = [
        'title', 
        'operation', 
        'line_name',
        'duration_display', 
        'has_test',
        'completion_count',
        'is_active'
    ]
    list_filter = ['operation__line__unit', 'operation__line', 'is_active', 'has_voice']
    search_fields = ['title', 'operation__name', 'operation__line__name']
    list_editable = ['is_active']
    inlines = [MCQTestInline]
    
    fieldsets = (
        ('Video Details', {
            'fields': (
                'operation',
                'title', 
                'description',
                ('has_voice', 'has_callouts'),
                'order'
            )
        }),
        ('Media', {
            'fields': ('video_file', 'thumbnail', 'preview_thumbnail', 'duration_seconds')
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
    )
    
    readonly_fields = ['preview_thumbnail']
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'operation__line__unit'
        ).annotate(
            _completion_count=Count('completions', filter=Q(completions__is_completed=True))
        ).prefetch_related('mcq_test')
    
    def line_name(self, obj):
        return obj.operation.line.name
    line_name.short_description = 'Line'
    
    def duration_display(self, obj):
        minutes = obj.duration_seconds // 60
        seconds = obj.duration_seconds % 60
        return f"{minutes}:{seconds:02d}"
    duration_display.short_description = 'Duration'
    
    def has_test(self, obj):
        try:
            test = obj.mcq_test
            return format_html(
                '✅ {} questions',
                test.questions.count()
            )
        except MCQTest.DoesNotExist:
            return format_html('<span style="color: #dc2626;">❌ No test</span>')
    has_test.short_description = 'MCQ Test'
    
    def completion_count(self, obj):
        return obj._completion_count
    completion_count.short_description = 'Completions'
    
    def preview_thumbnail(self, obj):
        if obj.thumbnail:
            return format_html('<img src="{}" style="max-height: 150px; border-radius: 8px;"/>', obj.thumbnail.url)
        return "No thumbnail"
    preview_thumbnail.short_description = 'Preview'


@admin.register(MCQTest)
class MCQTestAdmin(admin.ModelAdmin):
    """Manage tests with inline questions"""
    list_display = [
        'title', 
        'video_title',
        'question_count', 
        'passing_score', 
        'time_limit_minutes',
        'attempt_count',
        'avg_score',
        'is_active'
    ]
    list_filter = ['is_active', 'video__operation__line__unit']
    search_fields = ['title', 'video__title']
    inlines = [QuestionInline]
    
    fieldsets = (
        ('Test Configuration', {
            'fields': ('video', 'title', 'description', 'passing_score', 'time_limit_minutes', 'is_active')
        }),
    )
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('video').annotate(
            _question_count=Count('questions'),
            _attempt_count=Count('attempts'),
            _avg_score=Avg('attempts__score', filter=Q(attempts__status='completed'))
        )
    
    def video_title(self, obj):
        return obj.video.title
    video_title.short_description = 'Video'
    
    def question_count(self, obj):
        return obj._question_count
    question_count.short_description = 'Questions'
    
    def attempt_count(self, obj):
        return obj._attempt_count
    attempt_count.short_description = 'Attempts'
    
    def avg_score(self, obj):
        if obj._avg_score:
            return f"{obj._avg_score:.1f}%"
        return "N/A"
    avg_score.short_description = 'Avg Score'


# ============================================================================
# USER MANAGEMENT
# ============================================================================

@admin.register(EmployeeProfile)
class EmployeeProfileAdmin(admin.ModelAdmin):
    list_display = [
        'employee_id', 
        'user_full_name', 
        'plant', 
        'unit', 
        'department',
        'video_completions',
        'test_attempts'
    ]
    list_filter = ['plant', 'unit', 'department']
    search_fields = ['employee_id', 'user__username', 'user__first_name', 'user__last_name']
    
    fieldsets = (
        ('User Account', {
            'fields': ('user', 'employee_id', 'biometric_id')
        }),
        ('Organization', {
            'fields': ('plant', 'unit', 'department')
        }),
    )
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('user').annotate(
            _video_count=Count('user__video_completions', filter=Q(user__video_completions__is_completed=True)),
            _test_count=Count('user__test_attempts', filter=Q(user__test_attempts__status='completed'))
        )
    
    def user_full_name(self, obj):
        return obj.user.get_full_name() or obj.user.username
    user_full_name.short_description = 'Name'
    
    def video_completions(self, obj):
        return obj._video_count
    video_completions.short_description = 'Videos Completed'
    
    def test_attempts(self, obj):
        return obj._test_count
    test_attempts.short_description = 'Tests Completed'


# ============================================================================
# DASHBOARD & REPORTS
# ============================================================================

class DashboardAdmin(admin.ModelAdmin):
    """Custom admin for consolidated dashboard, guide, and reports"""
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('dashboard/', self.admin_site.admin_view(self.dashboard_view), name='training_dashboard'),
            path('guide/', self.admin_site.admin_view(self.guide_view), name='training_guide'),
            path('export-video-report/', self.admin_site.admin_view(self.export_video_report), name='export_video_report'),
            path('export-test-report/', self.admin_site.admin_view(self.export_test_report), name='export_test_report'),
            path('export-employee-report/', self.admin_site.admin_view(self.export_employee_report), name='export_employee_report'),
        ]
        return custom_urls + urls
    
    def guide_view(self, request):
        """Comprehensive admin guide page"""
        context = {
            'title': 'Admin Guide',
            'site_title': 'Process DOJO Guide',
        }
        return render(request, 'admin/training_guide.html', context)
    
    def dashboard_view(self, request):
        """Single consolidated training dashboard with all stats"""
        from django.db.models import F, Case, When, IntegerField
        
        # Overall statistics
        total_employees = EmployeeProfile.objects.count()
        total_videos = TrainingVideo.objects.filter(is_active=True).count()
        total_tests = MCQTest.objects.filter(is_active=True).count()
        
        # Video completion summary
        video_completions = VideoCompletion.objects.all()
        video_completion_rate = video_completions.filter(
            is_completed=True
        ).count() / max(video_completions.count(), 1) * 100
        
        # Test pass rate
        test_attempts = TestAttempt.objects.filter(status='completed')
        test_pass_rate = test_attempts.filter(passed=True).count() / max(test_attempts.count(), 1) * 100
        
        # Top performers
        top_performers = EmployeeProfile.objects.annotate(
            videos_completed=Count('user__video_completions', filter=Q(user__video_completions__is_completed=True)),
            tests_passed=Count('user__test_attempts', filter=Q(user__test_attempts__passed=True)),
            avg_score=Avg('user__test_attempts__score', filter=Q(user__test_attempts__status='completed'))
        ).filter(videos_completed__gt=0).order_by('-videos_completed', '-tests_passed')[:10]
        
        # Videos needing attention
        videos_needing_attention = TrainingVideo.objects.annotate(
            completion_count=Count('completions', filter=Q(completions__is_completed=True)),
            has_test=Count('mcq_test')
        ).filter(
            Q(completion_count__lt=5) | Q(has_test=0),
            is_active=True
        ).select_related('operation__line__unit')[:15]
        
        # Problem tests
        problem_tests = MCQTest.objects.annotate(
            total_attempts=Count('attempts', filter=Q(attempts__status='completed')),
            passed_attempts=Count('attempts', filter=Q(attempts__passed=True)),
            avg_score=Avg('attempts__score', filter=Q(attempts__status='completed'))
        ).filter(
            total_attempts__gte=3
        ).annotate(
            pass_rate=Case(
                When(total_attempts=0, then=0),
                default=F('passed_attempts') * 100 / F('total_attempts'),
                output_field=IntegerField()
            )
        ).filter(pass_rate__lt=70).select_related('video')[:10]
        
        # Unit progress
        unit_progress = Unit.objects.annotate(
            total_videos=Count('lines__operations__videos', filter=Q(lines__operations__videos__is_active=True)),
            total_completions=Count('lines__operations__videos__completions', 
                                   filter=Q(lines__operations__videos__completions__is_completed=True)),
            total_tests=Count('lines__operations__videos__mcq_test', 
                            filter=Q(lines__operations__videos__mcq_test__is_active=True)),
            total_test_attempts=Count('lines__operations__videos__mcq_test__attempts',
                                     filter=Q(lines__operations__videos__mcq_test__attempts__status='completed'))
        ).filter(is_active=True)
        
        # Recent activity (last 7 days)
        seven_days_ago = timezone.now() - timedelta(days=7)
        
        recent_completions = VideoCompletion.objects.filter(
            is_completed=True,
            last_watched_at__gte=seven_days_ago
        ).select_related('user__employee_profile', 'video__operation').order_by('-last_watched_at')[:15]
        
        recent_tests = TestAttempt.objects.filter(
            status='completed',
            completed_at__gte=seven_days_ago
        ).select_related('user__employee_profile', 'test__video').order_by('-completed_at')[:15]
        
        context = {
            'total_employees': total_employees,
            'total_videos': total_videos,
            'total_tests': total_tests,
            'video_completion_rate': round(video_completion_rate, 1),
            'test_pass_rate': round(test_pass_rate, 1),
            'top_performers': top_performers,
            'videos_needing_attention': videos_needing_attention,
            'problem_tests': problem_tests,
            'unit_progress': unit_progress,
            'recent_completions': recent_completions,
            'recent_tests': recent_tests,
            'title': 'Training Dashboard',
            'site_title': 'Process DOJO Dashboard',
        }
        
        return render(request, 'admin/training_dashboard.html', context)
    
    def export_video_report(self, request):
        """Export comprehensive video completion report as styled Excel"""
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Video Completions"
        
        # Define styles
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        title_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_font = Font(color="FFFFFF", bold=True, size=16)
        
        border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # Add title
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = '📹 Process DOJO - Video Completion Report'
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 30
        
        # Add metadata
        ws.merge_cells('A2:K2')
        meta_cell = ws['A2']
        meta_cell.value = f'Generated: {datetime.now().strftime("%B %d, %Y at %H:%M:%S")}'
        meta_cell.font = Font(italic=True, size=10, color='64748B')
        meta_cell.alignment = center_align
        
        # Add empty row
        ws.row_dimensions[3].height = 5
        
        # Add headers
        headers = [
            'Employee ID', 'Employee Name', 'Unit', 'Department',
            'Video Title', 'Operation', 'Line',
            'Completion %', 'Status', 'Access Count', 'Last Watched'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        ws.row_dimensions[4].height = 25
        
        # Get data
        completions = VideoCompletion.objects.select_related(
            'user__employee_profile',
            'video__operation__line__unit'
        ).order_by('user__employee_profile__employee_id', '-last_watched_at')
        
        # Add data rows
        row_num = 5
        for comp in completions:
            try:
                profile = comp.user.employee_profile
                
                # Data
                ws.cell(row=row_num, column=1, value=profile.employee_id)
                ws.cell(row=row_num, column=2, value=comp.user.get_full_name() or comp.user.username)
                ws.cell(row=row_num, column=3, value=profile.unit)
                ws.cell(row=row_num, column=4, value=profile.department)
                ws.cell(row=row_num, column=5, value=comp.video.title)
                ws.cell(row=row_num, column=6, value=comp.video.operation.name)
                ws.cell(row=row_num, column=7, value=comp.video.operation.line.name)
                
                # Completion percentage
                perc_cell = ws.cell(row=row_num, column=8, value=comp.completion_percentage)
                perc_cell.number_format = '0.0"%"'
                perc_cell.alignment = center_align
                
                # Status with color
                status_cell = ws.cell(row=row_num, column=9, value='✓ Completed' if comp.is_completed else 'In Progress')
                status_cell.alignment = center_align
                if comp.is_completed:
                    status_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    status_cell.font = Font(color="065F46", bold=True)
                else:
                    status_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                    status_cell.font = Font(color="92400E")
                
                # Access count
                access_cell = ws.cell(row=row_num, column=10, value=comp.access_count)
                access_cell.alignment = center_align
                
                # Last watched
                ws.cell(row=row_num, column=11, value=comp.last_watched_at.strftime('%Y-%m-%d %H:%M'))
                
                # Apply borders
                for col in range(1, 12):
                    ws.cell(row=row_num, column=col).border = border
                
                # Alternate row colors
                if row_num % 2 == 0:
                    for col in range(1, 12):
                        cell = ws.cell(row=row_num, column=col)
                        if cell.fill.start_color.rgb in [None, '00000000']:  # Don't override status colors
                            cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                
                row_num += 1
            except:
                pass
        
        # Add summary row
        row_num += 1
        ws.merge_cells(f'A{row_num}:J{row_num}')
        summary_cell = ws.cell(row=row_num, column=1)
        summary_cell.value = f'Total Records: {completions.count()}'
        summary_cell.font = Font(bold=True, size=11)
        summary_cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        summary_cell.alignment = left_align
        
        # Adjust column widths
        column_widths = [15, 25, 15, 20, 35, 25, 20, 15, 15, 12, 18]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Freeze header rows
        ws.freeze_panes = 'A5'
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="video_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    def export_test_report(self, request):
        """Export comprehensive test performance report as styled Excel"""
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Performance"
        
        # Define styles
        header_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        title_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_font = Font(color="FFFFFF", bold=True, size=16)
        
        border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # Add title
        ws.merge_cells('A1:P1')
        title_cell = ws['A1']
        title_cell.value = '📝 Process DOJO - Test Performance Report'
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 30
        
        # Add metadata
        ws.merge_cells('A2:P2')
        meta_cell = ws['A2']
        meta_cell.value = f'Generated: {datetime.now().strftime("%B %d, %Y at %H:%M:%S")}'
        meta_cell.font = Font(italic=True, size=10, color='64748B')
        meta_cell.alignment = center_align
        
        # Add empty row
        ws.row_dimensions[3].height = 5
        
        # Add headers
        headers = [
            'Employee ID', 'Employee Name', 'Unit', 'Department',
            'Test Title', 'Video', 'Operation', 'Line',
            'Status', 'Score', 'Correct', 'Total Q', 'Result',
            'Started At', 'Completed At', 'Duration (min)'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        ws.row_dimensions[4].height = 25
        
        # Get data
        attempts = TestAttempt.objects.select_related(
            'user__employee_profile',
            'test__video__operation__line'
        ).order_by('user__employee_profile__employee_id', '-started_at')
        
        # Add data rows
        row_num = 5
        for attempt in attempts:
            try:
                profile = attempt.user.employee_profile
                duration = ''
                if attempt.completed_at:
                    duration = int((attempt.completed_at - attempt.started_at).total_seconds() / 60)
                
                # Basic data
                ws.cell(row=row_num, column=1, value=profile.employee_id)
                ws.cell(row=row_num, column=2, value=attempt.user.get_full_name() or attempt.user.username)
                ws.cell(row=row_num, column=3, value=profile.unit)
                ws.cell(row=row_num, column=4, value=profile.department)
                ws.cell(row=row_num, column=5, value=attempt.test.title)
                ws.cell(row=row_num, column=6, value=attempt.test.video.title)
                ws.cell(row=row_num, column=7, value=attempt.test.video.operation.name)
                ws.cell(row=row_num, column=8, value=attempt.test.video.operation.line.name)
                
                # Status
                status_cell = ws.cell(row=row_num, column=9, value=attempt.status.replace('_', ' ').title())
                status_cell.alignment = center_align
                if attempt.status == 'completed':
                    status_cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                    status_cell.font = Font(color="1E40AF")
                
                # Score with formatting
                score_cell = ws.cell(row=row_num, column=10, value=attempt.score)
                score_cell.number_format = '0.0"%"'
                score_cell.alignment = center_align
                
                # Score color coding
                if attempt.score >= 80:
                    score_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    score_cell.font = Font(color="065F46", bold=True)
                elif attempt.score >= 60:
                    score_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                    score_cell.font = Font(color="92400E")
                else:
                    score_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    score_cell.font = Font(color="991B1B", bold=True)
                
                # Answers
                correct_cell = ws.cell(row=row_num, column=11, value=attempt.correct_answers)
                correct_cell.alignment = center_align
                
                total_cell = ws.cell(row=row_num, column=12, value=attempt.total_questions)
                total_cell.alignment = center_align
                
                # Pass/Fail
                result_cell = ws.cell(row=row_num, column=13, value='✓ PASSED' if attempt.passed else '✗ FAILED')
                result_cell.alignment = center_align
                if attempt.passed:
                    result_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    result_cell.font = Font(color="065F46", bold=True)
                else:
                    result_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    result_cell.font = Font(color="991B1B", bold=True)
                
                # Timestamps
                ws.cell(row=row_num, column=14, value=attempt.started_at.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row_num, column=15, value=attempt.completed_at.strftime('%Y-%m-%d %H:%M') if attempt.completed_at else 'In Progress')
                
                duration_cell = ws.cell(row=row_num, column=16, value=duration if duration else 'N/A')
                duration_cell.alignment = center_align
                
                # Apply borders
                for col in range(1, 17):
                    ws.cell(row=row_num, column=col).border = border
                
                # Alternate row colors
                if row_num % 2 == 0:
                    for col in range(1, 17):
                        cell = ws.cell(row=row_num, column=col)
                        if cell.fill.start_color.rgb in [None, '00000000']:
                            cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                
                row_num += 1
            except:
                pass
        
        # Add summary
        row_num += 1
        
        # Statistics
        total_attempts = attempts.count()
        completed = attempts.filter(status='completed').count()
        passed = attempts.filter(passed=True).count()
        avg_score = attempts.filter(status='completed').aggregate(Avg('score'))['score__avg'] or 0
        
        ws.merge_cells(f'A{row_num}:P{row_num}')
        summary_cell = ws.cell(row=row_num, column=1)
        summary_cell.value = f'Summary: {total_attempts} Total Attempts | {completed} Completed | {passed} Passed | Avg Score: {avg_score:.1f}%'
        summary_cell.font = Font(bold=True, size=11)
        summary_cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        summary_cell.alignment = center_align
        
        # Adjust column widths
        column_widths = [12, 22, 12, 18, 28, 30, 22, 18, 12, 10, 8, 8, 12, 16, 16, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Freeze header rows
        ws.freeze_panes = 'A5'
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="test_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    def export_employee_report(self, request):
        """Export employee training progress summary as styled Excel"""
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Summary"
        
        # Define styles
        header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        title_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_font = Font(color="FFFFFF", bold=True, size=16)
        
        border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # Add title
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = '👥 Process DOJO - Employee Training Summary'
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 30
        
        # Add metadata
        ws.merge_cells('A2:N2')
        meta_cell = ws['A2']
        meta_cell.value = f'Generated: {datetime.now().strftime("%B %d, %Y at %H:%M:%S")}'
        meta_cell.font = Font(italic=True, size=10, color='64748B')
        meta_cell.alignment = center_align
        
        # Add empty row
        ws.row_dimensions[3].height = 5
        
        # Add headers
        headers = [
            'Employee ID', 'Employee Name', 'Plant', 'Unit', 'Department',
            'Videos\nCompleted', 'Videos\nIn Progress', 'Total Video\nAccess',
            'Tests\nCompleted', 'Tests\nPassed', 'Tests\nFailed', 
            'Avg Test\nScore', 'Total\nLogins', 'Last Login'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        ws.row_dimensions[4].height = 35
        
        # Get data
        employees = EmployeeProfile.objects.annotate(
            videos_completed=Count('user__video_completions', filter=Q(user__video_completions__is_completed=True)),
            videos_in_progress=Count('user__video_completions', 
                                    filter=Q(user__video_completions__is_completed=False, 
                                           user__video_completions__completion_percentage__gt=0)),
            total_video_access=Count('user__video_completions'),
            tests_completed=Count('user__test_attempts', filter=Q(user__test_attempts__status='completed')),
            tests_passed=Count('user__test_attempts', filter=Q(user__test_attempts__passed=True)),
            tests_failed=Count('user__test_attempts', 
                             filter=Q(user__test_attempts__status='completed', user__test_attempts__passed=False)),
            avg_score=Avg('user__test_attempts__score', filter=Q(user__test_attempts__status='completed')),
            total_logins=Count('user__login_sessions')
        ).select_related('user')
        
        # Add data rows
        row_num = 5
        for emp in employees:
            last_login = LoginSession.objects.filter(user=emp.user).order_by('-login_time').first()
            
            # Basic info
            ws.cell(row=row_num, column=1, value=emp.employee_id)
            ws.cell(row=row_num, column=2, value=emp.user.get_full_name() or emp.user.username)
            ws.cell(row=row_num, column=3, value=emp.plant)
            ws.cell(row=row_num, column=4, value=emp.unit)
            ws.cell(row=row_num, column=5, value=emp.department)
            
            # Videos completed (with color)
            vid_comp_cell = ws.cell(row=row_num, column=6, value=emp.videos_completed)
            vid_comp_cell.alignment = center_align
            if emp.videos_completed >= 10:
                vid_comp_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                vid_comp_cell.font = Font(color="065F46", bold=True)
            elif emp.videos_completed >= 5:
                vid_comp_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                vid_comp_cell.font = Font(color="92400E")
            
            # Videos in progress
            vid_prog_cell = ws.cell(row=row_num, column=7, value=emp.videos_in_progress)
            vid_prog_cell.alignment = center_align
            
            # Total access
            access_cell = ws.cell(row=row_num, column=8, value=emp.total_video_access)
            access_cell.alignment = center_align
            
            # Tests completed
            test_comp_cell = ws.cell(row=row_num, column=9, value=emp.tests_completed)
            test_comp_cell.alignment = center_align
            
            # Tests passed (with color)
            test_pass_cell = ws.cell(row=row_num, column=10, value=emp.tests_passed)
            test_pass_cell.alignment = center_align
            if emp.tests_passed >= 5:
                test_pass_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                test_pass_cell.font = Font(color="065F46", bold=True)
            
            # Tests failed (with color)
            test_fail_cell = ws.cell(row=row_num, column=11, value=emp.tests_failed)
            test_fail_cell.alignment = center_align
            if emp.tests_failed > 3:
                test_fail_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                test_fail_cell.font = Font(color="991B1B")
            
            # Average score (with color)
            if emp.avg_score:
                avg_cell = ws.cell(row=row_num, column=12, value=emp.avg_score)
                avg_cell.number_format = '0.0"%"'
                avg_cell.alignment = center_align
                
                if emp.avg_score >= 80:
                    avg_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    avg_cell.font = Font(color="065F46", bold=True)
                elif emp.avg_score >= 60:
                    avg_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                    avg_cell.font = Font(color="92400E")
                else:
                    avg_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    avg_cell.font = Font(color="991B1B")
            else:
                ws.cell(row=row_num, column=12, value='N/A').alignment = center_align
            
            # Login stats
            login_cell = ws.cell(row=row_num, column=13, value=emp.total_logins)
            login_cell.alignment = center_align
            
            ws.cell(row=row_num, column=14, value=last_login.login_time.strftime('%Y-%m-%d %H:%M') if last_login else 'Never')
            
            # Apply borders
            for col in range(1, 15):
                ws.cell(row=row_num, column=col).border = border
            
            # Alternate row colors
            if row_num % 2 == 0:
                for col in range(1, 15):
                    cell = ws.cell(row=row_num, column=col)
                    if cell.fill.start_color.rgb in [None, '00000000']:
                        cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            
            row_num += 1
        
        # Add summary
        row_num += 1
        
        # Statistics
        total_emp = employees.count()
        total_vid_comp = sum(e.videos_completed for e in employees)
        total_test_comp = sum(e.tests_completed for e in employees)
        active_emp = employees.filter(videos_completed__gt=0).count()
        
        ws.merge_cells(f'A{row_num}:N{row_num}')
        summary_cell = ws.cell(row=row_num, column=1)
        summary_cell.value = f'Summary: {total_emp} Employees | {active_emp} Active | {total_vid_comp} Video Completions | {total_test_comp} Tests Completed'
        summary_cell.font = Font(bold=True, size=11)
        summary_cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        summary_cell.alignment = center_align
        
        # Adjust column widths
        column_widths = [12, 22, 12, 12, 18, 10, 10, 10, 10, 10, 10, 10, 10, 16]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Freeze header rows
        ws.freeze_panes = 'A5'
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="employee_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response


# Register dashboard
dashboard_admin = DashboardAdmin(Unit, admin.site)

# Add dashboard URLs to admin site
original_get_urls = admin.site.get_urls

def get_urls():
    urls = original_get_urls()
    custom_urls = [
        path('dashboard/', admin.site.admin_view(dashboard_admin.dashboard_view), name='training_dashboard'),
        path('guide/', admin.site.admin_view(dashboard_admin.guide_view), name='training_guide'),
        path('export-video-report/', admin.site.admin_view(dashboard_admin.export_video_report), name='export_video_report'),
        path('export-test-report/', admin.site.admin_view(dashboard_admin.export_test_report), name='export_test_report'),
        path('export-employee-report/', admin.site.admin_view(dashboard_admin.export_employee_report), name='export_employee_report'),
    ]
    return custom_urls + urls

admin.site.get_urls = get_urls


# Register tracking admins (read-only)
@admin.register(VideoCompletion)
class VideoCompletionAdmin(admin.ModelAdmin):
    list_display = ['user', 'employee_id', 'video', 'completion_percentage', 'is_completed', 'last_watched_at']
    list_filter = ['is_completed', 'video__operation__line__unit']
    search_fields = ['user__username', 'user__employee_profile__employee_id', 'video__title']
    date_hierarchy = 'last_watched_at'
    
    def has_add_permission(self, request):
        return False
    
    def employee_id(self, obj):
        try:
            return obj.user.employee_profile.employee_id
        except:
            return 'N/A'
    employee_id.short_description = 'Employee ID'


@admin.register(TestAttempt)
class TestAttemptAdmin(admin.ModelAdmin):
    list_display = ['user', 'employee_id', 'test', 'status', 'score', 'passed', 'started_at', 'completed_at']
    list_filter = ['status', 'passed', 'test__video__operation__line__unit']
    search_fields = ['user__username', 'user__employee_profile__employee_id', 'test__title']
    date_hierarchy = 'started_at'
    
    def has_add_permission(self, request):
        return False
    
    def employee_id(self, obj):
        try:
            return obj.user.employee_profile.employee_id
        except:
            return 'N/A'
    employee_id.short_description = 'Employee ID'


@admin.register(LoginSession)
class LoginSessionAdmin(admin.ModelAdmin):
    list_display = ['user', 'employee_id', 'login_time', 'logout_time', 'session_duration_minutes']
    list_filter = ['login_time']
    search_fields = ['user__username', 'user__employee_profile__employee_id']
    date_hierarchy = 'login_time'
    
    def has_add_permission(self, request):
        return False
    
    def employee_id(self, obj):
        try:
            return obj.user.employee_profile.employee_id
        except:
            return 'N/A'
    employee_id.short_description = 'Employee ID'


# Customize admin site branding
admin.site.site_header = "Renata AI - Training Management"
admin.site.site_title = "Renata AI"
admin.site.index_title = "Training Management System"
admin.site.index_template = 'admin/dashboard_index.html'